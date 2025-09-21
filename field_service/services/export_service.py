from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from typing import Iterable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from field_service.db import models as m
from field_service.db.session import SessionLocal
from field_service.services.settings_service import get_timezone

UTC = timezone.utc


@dataclass(slots=True)
class ExportBundle:
    csv_filename: str
    csv_bytes: bytes
    xlsx_filename: str
    xlsx_bytes: bytes


def _ensure_utc(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def _format_dt(value: Optional[datetime]) -> str:
    if value is None:
        return ""
    return _ensure_utc(value).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _format_decimal(value: Optional[Decimal | float]) -> str:
    if value is None:
        return "0.00"
    quantized = Decimal(value).quantize(Decimal("0.01"))
    return f"{quantized:.2f}"


def _format_bool(value: Optional[bool]) -> str:
    return "true" if value else "false"


def _render_csv(columns: Sequence[str], rows: Sequence[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, delimiter=";")
    writer.writeheader()
    for row in rows:
        writer.writerow({col: row.get(col, "") for col in columns})
    return buffer.getvalue().encode("utf-8-sig")


def _render_xlsx(
    sheet_name: str, columns: Sequence[str], rows: Sequence[dict[str, str]]
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(list(columns))
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    for idx, column in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(column) + 2)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _make_bundle(
    prefix: str, columns: Sequence[str], rows: Sequence[dict[str, str]]
) -> ExportBundle:
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    csv_bytes = _render_csv(columns, rows)
    xlsx_bytes = _render_xlsx(prefix, columns, rows)
    return ExportBundle(
        csv_filename=f"{prefix}_{timestamp}.csv",
        csv_bytes=csv_bytes,
        xlsx_filename=f"{prefix}_{timestamp}.xlsx",
        xlsx_bytes=xlsx_bytes,
    )


async def export_orders(
    *,
    date_from: datetime,
    date_to: datetime,
    city_ids: Optional[Iterable[int]] = None,
) -> ExportBundle:
    columns = [
        "order_id",
        "created_at_utc",
        "closed_at_utc",
        "city",
        "district",
        "street",
        "house",
        "lat",
        "lon",
        "category",
        "status",
        "type",
        "timeslot_start_utc",
        "timeslot_end_utc",
        "late_visit",
        "company_payment",
        "total_sum",
        "user_name",
        "user_phone",
        "master_name",
        "master_phone",
        "cancel_reason",
    ]
    start_utc = _ensure_utc(date_from)
    end_utc = _ensure_utc(date_to) + timedelta(microseconds=1)
    tz = get_timezone()

    async with SessionLocal() as session:
        stmt = (
            select(
                m.orders.id.label("order_id"),
                m.orders.created_at.label("created_at"),
                m.orders.updated_at.label("updated_at"),
                m.cities.name.label("city"),
                m.districts.name.label("district"),
                m.streets.name.label("street"),
                m.orders.house.label("house"),
                m.orders.status.label("status"),
                m.orders.scheduled_date.label("scheduled_date"),
                m.orders.time_slot_start.label("slot_start"),
                m.orders.time_slot_end.label("slot_end"),
                m.orders.company_payment.label("company_payment"),
                m.orders.total_price.label("total_price"),
                m.orders.client_name.label("client_name"),
                m.orders.client_phone.label("client_phone"),
                assigned_master.full_name.label("master_name"),
                assigned_master.phone.label("master_phone"),
                (
                    select(func.max(m.order_status_history.created_at)).where(
                        (m.order_status_history.order_id == m.orders.id)
                        & (m.order_status_history.to_status == m.OrderStatus.CLOSED)
                    )
                )
                .scalar_subquery()
                .label("closed_at"),
                (
                    select(m.order_status_history.reason)
                    .where(
                        (m.order_status_history.order_id == m.orders.id)
                        & (m.order_status_history.to_status == m.OrderStatus.CANCELED)
                    )
                    .order_by(m.order_status_history.created_at.desc())
                    .limit(1)
                )
                .scalar_subquery()
                .label("cancel_reason"),
            )
            .join(m.cities, m.orders.city_id == m.cities.id)
            .join(m.districts, m.orders.district_id == m.districts.id, isouter=True)
            .join(m.streets, m.orders.street_id == m.streets.id, isouter=True)
            .join(
                assigned_master,
                m.orders.assigned_master_id == assigned_master.id,
                isouter=True,
            )
            .where(m.orders.created_at >= start_utc, m.orders.created_at <= end_utc)
            .order_by(m.orders.created_at)
        )
        if city_ids:
            stmt = stmt.where(m.orders.city_id.in_(list(city_ids)))
        result = await session.execute(stmt)
        rows = []
        for row in result:
            scheduled_date = row.scheduled_date
            slot_start = row.slot_start
            slot_end = row.slot_end
            ts_start = (
                datetime.combine(scheduled_date, slot_start, tzinfo=tz)
                if scheduled_date and slot_start
                else None
            )
            ts_end = (
                datetime.combine(scheduled_date, slot_end, tzinfo=tz)
                if scheduled_date and slot_end
                else None
            )
            closed_at = row.closed_at or row.updated_at
            late_visit = (
                ts_end is not None
                and closed_at is not None
                and _ensure_utc(closed_at) > _ensure_utc(ts_end)
            )
            rows.append(
                {
                    "order_id": str(row.order_id),
                    "created_at_utc": _format_dt(row.created_at),
                    "closed_at_utc": _format_dt(row.closed_at),
                    "city": row.city or "",
                    "district": row.district or "",
                    "street": row.street or "",
                    "house": row.house or "",
                    "lat": "",
                    "lon": "",
                    "category": "",
                    "status": str(row.status),
                    "type": (
                        "GUARANTEE"
                        if str(row.status) == m.OrderStatus.GUARANTEE.value
                        else "NORMAL"
                    ),
                    "timeslot_start_utc": _format_dt(ts_start),
                    "timeslot_end_utc": _format_dt(ts_end),
                    "late_visit": _format_bool(late_visit),
                    "company_payment": _format_decimal(row.company_payment),
                    "total_sum": _format_decimal(row.total_price),
                    "user_name": row.client_name or "",
                    "user_phone": row.client_phone or "",
                    "master_name": row.master_name or "",
                    "master_phone": row.master_phone or "",
                    "cancel_reason": row.cancel_reason or "",
                }
            )
    return _make_bundle("orders", columns, rows)


async def export_commissions(
    *,
    date_from: datetime,
    date_to: datetime,
    city_ids: Optional[Iterable[int]] = None,
) -> ExportBundle:
    columns = [
        "commission_id",
        "order_id",
        "master_id",
        "master_name",
        "master_phone",
        "amount",
        "rate",
        "created_at_utc",
        "deadline_at_utc",
        "paid_reported_at_utc",
        "paid_approved_at_utc",
        "paid_amount",
        "is_paid",
        "has_checks",
        "snapshot_methods",
        "snapshot_card_number_last4",
        "snapshot_sbp_phone_masked",
    ]
    start_utc = _ensure_utc(date_from)
    end_utc = _ensure_utc(date_to) + timedelta(microseconds=1)

    async with SessionLocal() as session:
        master_alias = m.masters.alias("commission_master")
        stmt = (
            select(
                m.commissions.id,
                m.commissions.order_id,
                m.commissions.master_id,
                master_alias.full_name,
                master_alias.phone,
                m.commissions.amount,
                m.commissions.rate,
                m.commissions.created_at,
                m.commissions.deadline_at,
                m.commissions.paid_reported_at,
                m.commissions.paid_approved_at,
                m.commissions.paid_amount,
                m.commissions.is_paid,
                m.commissions.has_checks,
                m.commissions.pay_to_snapshot,
            )
            .join(master_alias, master_alias.id == m.commissions.master_id)
            .join(m.orders, m.orders.id == m.commissions.order_id)
            .where(
                m.commissions.created_at >= start_utc,
                m.commissions.created_at <= end_utc,
            )
            .order_by(m.commissions.created_at)
        )
        if city_ids:
            stmt = stmt.where(m.orders.city_id.in_(list(city_ids)))
        result = await session.execute(stmt)
        rows = []
        for row in result:
            snapshot = (
                row.pay_to_snapshot if isinstance(row.pay_to_snapshot, dict) else {}
            )
            methods = snapshot.get("methods")
            if isinstance(methods, list):
                snapshot_methods = ",".join(str(meth) for meth in methods)
            elif methods:
                snapshot_methods = str(methods)
            else:
                snapshot_methods = ""
            rows.append(
                {
                    "commission_id": str(row.id),
                    "order_id": str(row.order_id),
                    "master_id": str(row.master_id),
                    "master_name": row.full_name or "",
                    "master_phone": row.phone or "",
                    "amount": _format_decimal(row.amount),
                    "rate": _format_decimal(row.rate),
                    "created_at_utc": _format_dt(row.created_at),
                    "deadline_at_utc": _format_dt(row.deadline_at),
                    "paid_reported_at_utc": _format_dt(row.paid_reported_at),
                    "paid_approved_at_utc": _format_dt(row.paid_approved_at),
                    "paid_amount": _format_decimal(row.paid_amount),
                    "is_paid": _format_bool(row.is_paid),
                    "has_checks": _format_bool(row.has_checks),
                    "snapshot_methods": snapshot_methods,
                    "snapshot_card_number_last4": snapshot.get("card_number_last4", "")
                    or "",
                    "snapshot_sbp_phone_masked": snapshot.get("sbp_phone_masked", "")
                    or "",
                }
            )
    return _make_bundle("commissions", columns, rows)


async def export_referral_rewards(
    *,
    date_from: datetime,
    date_to: datetime,
    city_ids: Optional[Iterable[int]] = None,
) -> ExportBundle:
    columns = [
        "reward_id",
        "master_id",
        "order_id",
        "level",
        "amount",
        "created_at_utc",
    ]
    start_utc = _ensure_utc(date_from)
    end_utc = _ensure_utc(date_to) + timedelta(microseconds=1)

    async with SessionLocal() as session:
        stmt = (
            select(
                m.referral_rewards.id,
                m.referral_rewards.referrer_id,
                m.referral_rewards.commission_id,
                m.referral_rewards.level,
                m.referral_rewards.amount,
                m.referral_rewards.created_at,
                m.orders.city_id,
            )
            .join(m.commissions, m.commissions.id == m.referral_rewards.commission_id)
            .join(m.orders, m.orders.id == m.commissions.order_id)
            .where(
                m.referral_rewards.created_at >= start_utc,
                m.referral_rewards.created_at <= end_utc,
            )
            .order_by(m.referral_rewards.created_at)
        )
        if city_ids:
            stmt = stmt.where(m.orders.city_id.in_(list(city_ids)))
        result = await session.execute(stmt)
        rows = [
            {
                "reward_id": str(row.id),
                "master_id": str(row.referrer_id),
                "order_id": str(row.commission_id),
                "level": str(row.level),
                "amount": _format_decimal(row.amount),
                "created_at_utc": _format_dt(row.created_at),
            }
            for row in result
        ]
    return _make_bundle("ref_rewards", columns, rows)
