from __future__ import annotations

from datetime import date, datetime, time, timedelta, timezone
from decimal import Decimal
from zoneinfo import ZoneInfo
import io

import pytest
from openpyxl import load_workbook

from field_service.db import models as m
from field_service.services import export_service


# Tests for date/datetime compatibility
@pytest.mark.asyncio
async def test_export_orders_with_date_objects(monkeypatch, async_session):
    """Test that export_orders works with date objects (not datetime)."""
    monkeypatch.setattr(export_service, "get_timezone", lambda: ZoneInfo("UTC"))

    city = m.cities(name="TestCity")
    async_session.add(city)
    await async_session.flush()

    order = m.orders(
        city_id=city.id,
        status=m.OrderStatus.NEW,
        type=m.OrderType.NORMAL,
        total_sum=Decimal("1000.00"),
        created_at=datetime(2025, 1, 15, 10, 30, tzinfo=timezone.utc),
        description="Test",
    )
    async_session.add(order)
    await async_session.flush()

    # Use date objects instead of datetime
    bundle = await export_service.export_orders(
        date_from=date(2025, 1, 15),
        date_to=date(2025, 1, 15),
        city_ids=[city.id],
        session=async_session,
    )

    assert bundle.csv_filename.startswith("orders_")
    csv_text = bundle.csv_bytes.decode("utf-8-sig")
    assert str(order.id) in csv_text


@pytest.mark.asyncio
async def test_export_commissions_with_date_objects(async_session):
    """Test that export_commissions works with date objects."""
    city = m.cities(name="TestCity")
    async_session.add(city)
    await async_session.flush()

    master = m.masters(
        full_name="Test Master",
        phone="+79991234567",
        city_id=city.id,
        verified=True,
        is_active=True,
    )
    async_session.add(master)
    await async_session.flush()

    order = m.orders(
        city_id=city.id,
        status=m.OrderStatus.PAYMENT,
        type=m.OrderType.NORMAL,
        total_sum=Decimal("2000.00"),
        assigned_master_id=master.id,
        created_at=datetime(2025, 1, 15, tzinfo=timezone.utc),
        description="Test",
    )
    async_session.add(order)
    await async_session.flush()

    commission = m.commissions(
        order_id=order.id,
        master_id=master.id,
        amount=Decimal("1000.00"),
        rate=Decimal("0.50"),
        status=m.CommissionStatus.WAIT_PAY,
        created_at=datetime(2025, 1, 15, 12, tzinfo=timezone.utc),
        deadline_at=datetime(2025, 1, 15, 15, tzinfo=timezone.utc),
        is_paid=False,
    )
    async_session.add(commission)
    await async_session.flush()

    # Use date objects
    bundle = await export_service.export_commissions(
        date_from=date(2025, 1, 15),
        date_to=date(2025, 1, 15),
        city_ids=[city.id],
        session=async_session,
    )

    csv_text = bundle.csv_bytes.decode("utf-8-sig")
    assert str(commission.id) in csv_text


@pytest.mark.asyncio
async def test_export_referral_rewards_with_date_objects(async_session):
    """Test that export_referral_rewards works with date objects."""
    city = m.cities(name="TestCity")
    async_session.add(city)
    await async_session.flush()

    referrer = m.masters(
        full_name="Referrer",
        phone="+79991111111",
        city_id=city.id,
        verified=True,
        is_active=True,
    )
    referred = m.masters(
        full_name="Referred",
        phone="+79992222222",
        city_id=city.id,
        verified=True,
        is_active=True,
    )
    async_session.add_all([referrer, referred])
    await async_session.flush()

    order = m.orders(
        city_id=city.id,
        status=m.OrderStatus.CLOSED,
        type=m.OrderType.NORMAL,
        assigned_master_id=referred.id,
        total_sum=Decimal("3000.00"),
        created_at=datetime(2025, 1, 15, tzinfo=timezone.utc),
    )
    async_session.add(order)
    await async_session.flush()

    commission = m.commissions(
        order_id=order.id,
        master_id=referred.id,
        amount=Decimal("1500.00"),
        rate=Decimal("0.50"),
        status=m.CommissionStatus.WAIT_PAY,
        created_at=datetime(2025, 1, 15, tzinfo=timezone.utc),
        deadline_at=datetime(2025, 1, 15, 3, tzinfo=timezone.utc),
        is_paid=False,
    )
    async_session.add(commission)
    await async_session.flush()

    reward = m.referral_rewards(
        referrer_id=referrer.id,
        referred_master_id=referred.id,
        commission_id=commission.id,
        level=1,
        percent=Decimal("10.00"),
        amount=Decimal("150.00"),
        status=m.ReferralRewardStatus.ACCRUED,
        created_at=datetime(2025, 1, 15, 14, tzinfo=timezone.utc),
    )
    async_session.add(reward)

    # Use date objects
    bundle = await export_service.export_referral_rewards(
        date_from=date(2025, 1, 15),
        date_to=date(2025, 1, 15),
        city_ids=[city.id],
        session=async_session,
    )

    csv_text = bundle.csv_bytes.decode("utf-8-sig")
    assert str(reward.id) in csv_text


@pytest.mark.asyncio
async def test_export_orders_bundle(monkeypatch, async_session):
    monkeypatch.setattr(export_service, "get_timezone", lambda: ZoneInfo("UTC"))

    city = m.cities(name=" ")
    async_session.add(city)
    await async_session.flush()

    district = m.districts(city_id=city.id, name="")
    street = m.streets(city_id=city.id, district_id=district.id, name="")
    master = m.masters(
        full_name=" ",
        phone="+79990001122",
        city_id=city.id,
        verified=True,
        is_active=True,
    )
    async_session.add_all([district, street, master])
    await async_session.flush()

    order = m.orders(
        city_id=city.id,
        district_id=district.id,
        street_id=street.id,
        house="10",
        lat=Decimal("55.123456"),
        lon=Decimal("37.654321"),
        category="ELECTRICS",
        status=m.OrderStatus.CLOSED,
        type=m.OrderType.NORMAL,
        late_visit=True,
        company_payment=Decimal("0"),
        total_sum=Decimal("3500.50"),
        client_name=" ",
        client_phone="+79991234567",
        assigned_master_id=master.id,
        created_at=datetime(2025, 9, 14, 12, tzinfo=timezone.utc),
        updated_at=datetime(2025, 9, 15, 12, tzinfo=timezone.utc),
        description="",
    )
    async_session.add(order)
    await async_session.flush()

    async_session.add(
        m.order_status_history(
            order_id=order.id,
            from_status=m.OrderStatus.WORKING,
            to_status=m.OrderStatus.CLOSED,
            created_at=datetime(2025, 9, 15, 11, tzinfo=timezone.utc),
        )
    )

    bundle = await export_service.export_orders(
        date_from=datetime(2025, 9, 14, tzinfo=timezone.utc),
        date_to=datetime(2025, 9, 16, tzinfo=timezone.utc),
        city_ids=[city.id],
        session=async_session,
    )

    assert bundle.csv_filename.startswith("orders_")
    assert bundle.xlsx_filename.startswith("orders_")

    csv_text = bundle.csv_bytes.decode("utf-8-sig").splitlines()
    header = csv_text[0].split(";")
    expected_columns = [
        "order_id",
        "created_at_utc",
        "closed_at_utc",
        "city",
        "district",
        "street",
        "house",
        "lat",
        "lon",
        "category",
        "status",
        "type",
        "timeslot_start_utc",
        "timeslot_end_utc",
        "late_visit",
        "company_payment",
        "total_sum",
        "user_name",
        "user_phone",
        "master_name",
        "master_phone",
        "cancel_reason",
    ]
    assert header == expected_columns

    values = dict(zip(header, csv_text[1].split(";")))
    assert values["city"] == " "
    assert values["district"] == ""
    assert values["street"] == ""
    assert values["house"] == "10"
    assert values["lat"] == "55.123456"
    assert values["lon"] == "37.654321"
    assert values["category"] == "ELECTRICS"
    assert values["status"] == "CLOSED"
    assert values["type"] == "NORMAL"
    assert values["late_visit"] == "true"
    assert values["company_payment"] == ""
    assert values["total_sum"] == "3500.50"
    assert values["user_name"] == " "
    assert values["user_phone"] == "+79991234567"
    assert values["master_name"] == " "
    assert values["master_phone"] == "+79990001122"
    assert values["timeslot_start_utc"] == "2025-09-15T10:00:00Z"
    assert values["timeslot_end_utc"] == "2025-09-15T13:00:00Z"

    wb = load_workbook(io.BytesIO(bundle.xlsx_bytes))
    assert wb.sheetnames == ["orders"]
    row = list(wb["orders"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_by_name = dict(zip(header, row))
    assert row_by_name["order_id"] == order.id
    assert row_by_name["total_sum"] == pytest.approx(3500.50)
    assert row_by_name["late_visit"] is True
    assert row_by_name["company_payment"] is None


@pytest.mark.asyncio
async def test_export_commissions(monkeypatch, async_session):
    city = m.cities(name=" ")
    async_session.add(city)
    await async_session.flush()

    master = m.masters(
        full_name="",
        phone="+79990002233",
        city_id=city.id,
        verified=True,
        is_active=True,
    )
    async_session.add(master)
    await async_session.flush()

    order = m.orders(
        city_id=city.id,
        status=m.OrderStatus.PAYMENT,
        type=m.OrderType.NORMAL,
        total_sum=Decimal("4000.00"),
        assigned_master_id=master.id,
        created_at=datetime(2025, 9, 10, tzinfo=timezone.utc),
        description="",
    )
    async_session.add(order)
    await async_session.flush()

    commission = m.commissions(
        order_id=order.id,
        master_id=master.id,
        amount=Decimal("2000.00"),
        rate=Decimal("0.50"),
        status=m.CommissionStatus.APPROVED,
        created_at=datetime(2025, 9, 11, tzinfo=timezone.utc),
        deadline_at=datetime(2025, 9, 11, 3, tzinfo=timezone.utc),
        paid_reported_at=datetime(2025, 9, 11, 1, tzinfo=timezone.utc),
        paid_approved_at=datetime(2025, 9, 11, 2, tzinfo=timezone.utc),
        paid_amount=Decimal("2000.00"),
        is_paid=True,
        pay_to_snapshot={
            "methods": ["card", "sbp"],
            "card_number_last4": "4242",
            "sbp_phone_masked": "+71234",
        },
    )
    async_session.add(commission)
    await async_session.flush()

    async_session.add(
        m.attachments(
            entity_type=m.AttachmentEntity.COMMISSION,
            entity_id=commission.id,
            file_type=m.AttachmentFileType.PHOTO,
            file_id="file-check",
        )
    )

    bundle = await export_service.export_commissions(
        date_from=datetime(2025, 9, 10, tzinfo=timezone.utc),
        date_to=datetime(2025, 9, 12, tzinfo=timezone.utc),
        city_ids=[city.id],
        session=async_session,
    )

    rows = bundle.csv_bytes.decode("utf-8-sig").splitlines()
    header = rows[0].split(";")
    values = dict(zip(header, rows[1].split(";")))
    assert values["commission_id"] == str(commission.id)
    assert values["amount"] == "2000.00"
    assert values["rate"] == "0.50"
    assert values["is_paid"] == "true"
    assert values["has_checks"] == "true"
    assert values["snapshot_methods"] == "card,sbp"
    assert values["snapshot_card_number_last4"] == "4242"
    assert values["snapshot_sbp_phone_masked"] == "+71234"

    wb = load_workbook(io.BytesIO(bundle.xlsx_bytes))
    assert wb.sheetnames == ["commissions"]
    data_row = list(wb["commissions"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_map = dict(zip(header, data_row))
    assert row_map["has_checks"] is True
    assert row_map["amount"] == pytest.approx(2000.00)


@pytest.mark.asyncio
async def test_export_referral_rewards(async_session):
    city = m.cities(name="")
    async_session.add(city)
    await async_session.flush()

    referrer = m.masters(full_name="", phone="+79990003344", city_id=city.id, verified=True, is_active=True)
    referred = m.masters(full_name="", phone="+79990004455", city_id=city.id, verified=True, is_active=True)
    async_session.add_all([referrer, referred])
    await async_session.flush()

    order = m.orders(
        city_id=city.id,
        status=m.OrderStatus.CLOSED,
        type=m.OrderType.NORMAL,
        assigned_master_id=referred.id,
        total_sum=Decimal("5000.00"),
        created_at=datetime(2025, 9, 10, tzinfo=timezone.utc),
    )
    async_session.add(order)
    await async_session.flush()

    commission = m.commissions(
        order_id=order.id,
        master_id=referred.id,
        amount=Decimal("2500.00"),
        rate=Decimal("0.50"),
        status=m.CommissionStatus.WAIT_PAY,
        created_at=datetime(2025, 9, 11, tzinfo=timezone.utc),
        deadline_at=datetime(2025, 9, 11, 3, tzinfo=timezone.utc),
        is_paid=False,
    )
    async_session.add(commission)
    await async_session.flush()

    reward = m.referral_rewards(
        referrer_id=referrer.id,
        referred_master_id=referred.id,
        commission_id=commission.id,
        level=1,
        percent=Decimal("10.00"),
        amount=Decimal("250.00"),
        status=m.ReferralRewardStatus.ACCRUED,
        created_at=datetime(2025, 9, 11, tzinfo=timezone.utc),
    )
    async_session.add(reward)

    bundle = await export_service.export_referral_rewards(
        date_from=datetime(2025, 9, 10, tzinfo=timezone.utc),
        date_to=datetime(2025, 9, 12, tzinfo=timezone.utc),
        city_ids=[city.id],
        session=async_session,
    )

    rows = bundle.csv_bytes.decode("utf-8-sig").splitlines()
    header = rows[0].split(";")
    assert header == [
        "reward_id",
        "master_id",
        "order_id",
        "commission_id",
        "level",
        "amount",
        "created_at_utc",
    ]
    values = dict(zip(header, rows[1].split(";")))
    assert values["commission_id"] == str(commission.id)
    assert values["master_id"] == str(referrer.id)
    assert values["order_id"] == str(order.id)
    assert values["amount"] == "250.00"

    wb = load_workbook(io.BytesIO(bundle.xlsx_bytes))
    assert wb.sheetnames == ["ref_rewards"]
    data_row = list(wb["ref_rewards"].iter_rows(min_row=2, max_row=2, values_only=True))[0]
    row_map = dict(zip(header, data_row))
    assert row_map["reward_id"] == reward.id
    assert row_map["level"] == reward.level
    assert row_map["amount"] == pytest.approx(250.00)

